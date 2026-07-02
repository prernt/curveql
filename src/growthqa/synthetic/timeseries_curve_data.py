#!/usr/bin/env python3
"""
timeseries_curve_data.py
------------------------
Synthetic growth-curve generator for GrowthQA.

The generator writes one wide CSV:

    timeseries_wide_<file-stem>.csv
        Columns:
            FileName, Test Id, Model Name, Is_Valid, Curve Subtype,
            T0.0 (h), T0.5 (h), ...

It also appends a dated record of every run to run_info.xlsx.

Design
------
The dataset has a fixed composition (see VALID_COMPOSITION / INVALID_COMPOSITION).
All curves are generated on a single time grid (default 0-16 h, 0.5 h step),
matching the 16 h training window used throughout GrowthQA.

Curve families
--------------
Valid curves use established single-phase growth-model forms:
    Logistic, Gompertz, ModifiedGompertz, Richards  (Zwietering et al. 1990; Kahm et al. 2010)
    decline = growth phase followed by a death phase (Monod 1949, four-phase structure)

Invalid curves cover both curves that single-phase fitting cannot represent and
recognised assay-failure modes:
    diauxic      - two sequential growth phases (Monod 1949); biologically real but
                   not representable by a single-phase model, so invalid for fitting
    obvious      - abrupt mid-curve crash or tail collapse (contamination / lysis)
    subtle       - localised reading artifact (bubble / condensation / misread)
    nearreal     - suppressed, non-stabilising plateau (insufficient stationary phase)
    decline_only - decay with no preceding growth (dead inoculum / wrong well)
    noise        - low signal-to-noise readings (no usable growth signal)
    nogrowth     - flat near-zero curve (negative control)

The four corruption families (obvious, subtle, nearreal, and the imperfections
injected into valid curves) are engineering constructions that mimic documented
plate-reader and culture failures.
"""

from __future__ import annotations

import argparse
import logging
import os
from datetime import datetime

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# 1) Growth-model definitions
# ---------------------------------------------------------------------------
# Logistic and Gompertz use the A, mu, lambda parameterisation.
# A = carrying capacity, mu = max growth rate, lambda = lag time.

def logistic(t, A, mu, lam):
    return A / (1.0 + np.exp((4.0 * mu / A) * (lam - t) + 2.0))


def gompertz(t, A, mu, lam):
    return A * np.exp(-np.exp((mu * np.e / A) * (lam - t) + 1.0))


def modified_gompertz(t, A, mu, lam, alpha, tshift):
    """Zwietering Gompertz with a bounded, saturating late adjustment.

    The primary term is the standard A, mu, lambda Gompertz (Zwietering et al. 1990),
    bounded by A. The secondary term adds a late adjustment of small amplitude
    ``alpha * A`` that switches on at ``tshift`` and saturates, modelling a mild
    drift of the stationary level without introducing a second growth phase. The
    curve is monotone non-decreasing, stays bounded by ``A * (1 + alpha)``, and
    reaches a finite plateau, so it remains a valid single-phase curve.
    """
    primary = A * np.exp(-np.exp((mu * np.e / A) * (lam - t) + 1.0))
    onset = np.clip(t - tshift, 0.0, None)
    secondary = alpha * A * (1.0 - np.exp(-onset))
    return primary + secondary


def richards(t, A, mu, lam, nu):
    return A * (1.0 + nu * np.exp((mu * (1.0 + nu) / A) * (lam - t))) ** (-1.0 / nu)


def diauxic(t, A1, mu1, lam1, A2, mu2, lam2):
    return logistic(t, A1, mu1, lam1) + logistic(t, A2, mu2, lam2)


def flat_line(t, baseline):
    return np.full_like(t, baseline)


# model name -> (function, parameter names)
MODEL_SPECS = {
    "Logistic": (logistic, ["A", "mu", "lam"]),
    "Gompertz": (gompertz, ["A", "mu", "lam"]),
    "ModifiedGompertz": (modified_gompertz, ["A", "mu", "lam", "alpha", "tshift"]),
    "Richards": (richards, ["A", "mu", "lam", "nu"]),
    "Diauxic": (diauxic, ["A1", "mu1", "lam1", "A2", "mu2", "lam2"]),
    "Flat": (flat_line, ["baseline"]),
}

# Sigmoidal models used for valid (and base-of-invalid) curves.
GROWTH_MODELS = {"Logistic", "Gompertz", "ModifiedGompertz", "Richards"}


# ---------------------------------------------------------------------------
# 2) Fixed dataset composition
# ---------------------------------------------------------------------------
# Explicit per-subtype counts are used instead of tunable weights so the
# dataset is reproducible and the exact composition can be reported.
# Lab data contributes only valid curves, so every invalid example here is
# the synthetic dataset's responsibility.

VALID_COMPOSITION = {
    "plain": 203,
    "fast": 100,
    "late": 100,
    "decline": 100,
}  # 503

INVALID_COMPOSITION = {
    "obvious": 80,
    "diauxic": 60,
    "subtle": 60,
    "nearreal": 55,
    "decline_only": 55,
    "noise": 45,
    "nogrowth": 42,
}  # 397

N_VALID = sum(VALID_COMPOSITION.values())
N_INVALID = sum(INVALID_COMPOSITION.values())
TARGET_CURVES = N_VALID + N_INVALID  # 900

# Imperfections injected into a subset of valid curves so the dataset contains
# clean valids, imperfect valids, and clearly invalid curves.
PCT_HIGH_QUALITY_VALID = 0.30   # fraction of valid curves generated with reduced noise
PCT_MISSING_CURVES = 0.10       # fraction of valid curves given some missing readings
MISSING_FRAC_PER_CURVE = 0.10
PCT_OUTLIER_CURVES = 0.05       # fraction of valid curves given negative blips
OUTLIER_FRAC_PER_CURVE = 0.05
OUTLIER_SCALE_MIN = 0.10
OUTLIER_SCALE_MAX = 0.30

# Decline (death-phase) parameters for the valid "decline" subtype.
VALID_DECLINE_K_MIN = 0.03
VALID_DECLINE_K_MAX = 0.15


# ---------------------------------------------------------------------------
# 3) Corruption / imperfection helpers
# ---------------------------------------------------------------------------

def inject_missing(y: pd.Series, frac: float, rng: np.random.Generator) -> pd.Series:
    if frac <= 0:
        return y
    n = len(y)
    k = max(1, int(round(frac * n)))
    idx = rng.choice(n, size=min(k, n), replace=False)
    y.iloc[idx] = np.nan
    return y


def inject_negative_outliers(
    y: pd.Series, frac: float, scale_min: float, scale_max: float, rng: np.random.Generator
) -> pd.Series:
    if frac <= 0:
        return y
    n = len(y)
    k = max(1, int(round(frac * n)))
    idx = rng.choice(n, size=min(k, n), replace=False)
    sub = rng.uniform(scale_min, scale_max, size=len(idx))
    y.iloc[idx] = y.iloc[idx].values - sub
    return y


def make_obvious_invalid(y: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """Abrupt mid-curve crash: drop to near-zero somewhere in the middle third."""
    y = np.asarray(y, dtype=float).copy()
    n = len(y)
    if n < 4:
        return y
    cut = rng.integers(n // 3, 2 * n // 3)
    drop_factor = rng.uniform(0.0, 0.2)
    y[cut:] = y[cut] * drop_factor
    return y


def make_sudden_collapse(y: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """Abrupt tail collapse: sharp drop late in the curve."""
    y = np.asarray(y, dtype=float).copy()
    n = len(y)
    if n < 4:
        return y
    cut = rng.integers(n // 2, n - 1)
    y[cut:] = y[cut:] * rng.uniform(0.05, 0.25)
    return y


def make_subtle_invalid(y: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """Localised dip (bubble / misreading) that does not look obviously wrong."""
    y = np.asarray(y, dtype=float).copy()
    n = len(y)
    if n < 6:
        return y
    center = rng.integers(n // 4, 3 * n // 4)
    width = rng.integers(2, min(6, n - center))
    max_y = np.nanmax(y) if np.any(np.isfinite(y)) else 1.0
    drop = rng.uniform(0.1, 0.4) * max_y
    y[center:center + width] = np.clip(y[center:center + width] - drop, 0.0, None)
    return y


def make_near_real_invalid(y: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """Realistic-looking but invalid: suppressed tail that never stabilises."""
    y = np.asarray(y, dtype=float).copy()
    n = len(y)
    if n < 6:
        return y
    start_tail = int(0.6 * n)
    factor = rng.uniform(0.4, 0.8)
    y[start_tail:] = y[start_tail:] * factor
    trend = np.linspace(0.0, rng.uniform(0.05, 0.15) * np.nanmax(y), n - start_tail)
    y[start_tail:] = np.clip(y[start_tail:] - trend, 0.0, None)
    return y


def make_decline_only_invalid(t: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """Starts high and decays with no preceding growth phase."""
    t = np.asarray(t, dtype=float)
    baseline = rng.uniform(0.4, 1.2)
    k = rng.uniform(0.05, 0.2)
    return baseline * np.exp(-k * (t - t.min()))


def make_noise_dominated(t: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """Low signal-to-noise readings with no usable growth signal."""
    t = np.asarray(t, dtype=float)
    baseline = rng.uniform(0.05, 0.2)
    y = baseline + rng.normal(0, 0.2, size=len(t))
    return np.clip(y, 0.0, None)


def apply_valid_decline(
    y: np.ndarray, t: np.ndarray, rng: np.random.Generator, t_start: float, k: float
) -> np.ndarray:
    """Structured death phase: exponential decay applied after a clear growth phase."""
    y = np.asarray(y, dtype=float).copy()
    decay = np.ones_like(t, dtype=float)
    mask = t > t_start
    decay[mask] = np.exp(-k * (t[mask] - t_start))
    return np.clip(y * decay, 0.0, None)


# ---------------------------------------------------------------------------
# 4) Curve sampling
# ---------------------------------------------------------------------------

def sample_valid_curve(
    subtype: str, time_points: np.ndarray, max_time: float, rng: np.random.Generator
) -> tuple[str, np.ndarray]:
    """Sample one valid single-phase growth curve for the given subtype."""
    model_name = rng.choice(sorted(GROWTH_MODELS))

    late_lam_range = (8.0, 12.0) if max_time <= 16.0 else (10.0, min(16.0, max_time))
    if subtype == "fast":
        mu_range, lam_range = (0.9, 1.5), (0.0, 2.0)
    elif subtype == "late":
        mu_range, lam_range = (0.8, 1.5), late_lam_range
    else:
        mu_range, lam_range = (0.2, 1.5), (0.0, min(10.0, max_time))

    pars = {"A": rng.uniform(0.5, 2.0), "mu": rng.uniform(*mu_range), "lam": rng.uniform(*lam_range)}
    if model_name == "ModifiedGompertz":
        pars.update({"alpha": rng.uniform(0.0, 0.15), "tshift": rng.uniform(0.4 * max_time, 0.7 * max_time)})
    if model_name == "Richards":
        pars.update({"nu": rng.uniform(0.5, 2.0)})
    return model_name, MODEL_SPECS[model_name][0](time_points, **pars)


def sample_diauxic_curve(
    time_points: np.ndarray, max_time: float, rng: np.random.Generator
) -> np.ndarray:
    """Diauxic growth is biologically genuine (Monod 1949) but contains two sequential
    growth phases. A single-phase parametric model, which is the GrowthQA / grofit
    fitting target, cannot represent it, so GrowthQA labels diauxic curves invalid.
    """
    pars = {
        "A1": rng.uniform(0.3, 1.0),
        "mu1": rng.uniform(0.2, 1.2),
        "lam1": rng.uniform(0.0, 5.0),
        "A2": rng.uniform(0.3, 1.0),
        "mu2": rng.uniform(0.2, 1.2),
        "lam2": rng.uniform(5.0, min(15.0, max_time)),
    }
    return diauxic(time_points, **pars)


def sample_invalid_curve(
    subtype: str, time_points: np.ndarray, max_time: float, rng: np.random.Generator
) -> tuple[str, np.ndarray]:
    """Sample one invalid curve for the given subtype. Returns (model_name, y)."""
    if subtype == "diauxic":
        return "Diauxic", sample_diauxic_curve(time_points, max_time, rng)

    if subtype in {"obvious", "subtle", "nearreal"}:
        base_model, base_curve = sample_valid_curve("plain", time_points, max_time, rng)
        if subtype == "obvious":
            # Merged failure family: mid-curve crash or tail collapse.
            y = make_obvious_invalid(base_curve, rng) if rng.random() < 0.5 \
                else make_sudden_collapse(base_curve, rng)
        elif subtype == "subtle":
            y = make_subtle_invalid(base_curve, rng)
        else:
            y = make_near_real_invalid(base_curve, rng)
        return base_model, y

    if subtype == "decline_only":
        return "Flat", make_decline_only_invalid(time_points, rng)
    if subtype == "noise":
        return "Flat", make_noise_dominated(time_points, rng)
    if subtype == "nogrowth":
        return "Flat", flat_line(time_points, rng.uniform(0.0, 0.1))
    raise ValueError(f"Unknown invalid subtype: {subtype}")


# ---------------------------------------------------------------------------
# 5) Dataset generation
# ---------------------------------------------------------------------------

def generate_synthetic_wide_df(
    *,
    tmax_hours: float = 16.0,
    step_hours: float = 0.5,
    seed: int = 123,
    noise_level: float = 0.05,
    file_stem: str = "syn",
    time_unit: str = "h",
) -> tuple[pd.DataFrame, dict]:
    """Generate the synthetic dataset as a single wide DataFrame.

    The composition is fixed by VALID_COMPOSITION / INVALID_COMPOSITION.
    Returns the DataFrame and a summary dict (counts per subtype).
    """
    rng = np.random.default_rng(seed)
    time_points = np.arange(0.0, tmax_hours + step_hours / 2.0, step_hours)

    plan: list[tuple[str, str]] = []
    for sub, n in VALID_COMPOSITION.items():
        plan.extend([("VALID", sub)] * n)
    for sub, n in INVALID_COMPOSITION.items():
        plan.extend([("INVALID", sub)] * n)
    rng.shuffle(plan)

    rows = []
    for i, (cls, subtype) in enumerate(plan, start=1):
        is_valid = cls == "VALID"

        if is_valid:
            model_name, y = sample_valid_curve(subtype, time_points, tmax_hours, rng)
            if subtype == "decline":
                if tmax_hours <= 16.0:
                    t_start = rng.uniform(8.0, 12.0)
                else:
                    t_start = rng.uniform(10.0, 14.0)
                k = rng.uniform(VALID_DECLINE_K_MIN, VALID_DECLINE_K_MAX)
                y = apply_valid_decline(y, time_points, rng, t_start, k)
        else:
            model_name, y = sample_invalid_curve(subtype, time_points, tmax_hours, rng)

        # Measurement noise. A fraction of valid curves are made cleaner.
        high_quality = is_valid and rng.random() < PCT_HIGH_QUALITY_VALID
        noise_std = noise_level * 0.3 if high_quality else noise_level
        y = y + rng.normal(0.0, noise_std, size=y.shape)

        # Imperfections only on valid curves (invalids are already corrupted).
        if is_valid:
            if rng.random() < PCT_MISSING_CURVES:
                y = inject_missing(pd.Series(y), MISSING_FRAC_PER_CURVE, rng).values
            if rng.random() < PCT_OUTLIER_CURVES:
                y = inject_negative_outliers(
                    pd.Series(y), OUTLIER_FRAC_PER_CURVE,
                    OUTLIER_SCALE_MIN, OUTLIER_SCALE_MAX, rng,
                ).values

        # OD600 is physically non-negative; clip after noise and injection.
        # Missing values (NaN) are preserved by np.clip.
        y = np.clip(y, 0.0, None)

        row = {
            "FileName": file_stem,
            "Test Id": f"{file_stem}_{i}",
            "Model Name": model_name,
            "Is_Valid": bool(is_valid),
            "Curve Subtype": subtype,
        }
        for j, t in enumerate(time_points):
            row[f"T{np.round(t, 6)} ({time_unit})"] = float(y[j])
        rows.append(row)

    df = pd.DataFrame(rows)

    valid_counts = {f"valid_{k}": v for k, v in VALID_COMPOSITION.items()}
    invalid_counts = {f"invalid_{k}": v for k, v in INVALID_COMPOSITION.items()}
    summary = {
        "tmax_hours": float(tmax_hours),
        "step_hours": float(step_hours),
        "seed": int(seed),
        "noise_level": float(noise_level),
        "n_curves": int(len(df)),
        "n_valid": int(N_VALID),
        "n_invalid": int(N_INVALID),
        "valid_counts": valid_counts,
        "invalid_counts": invalid_counts,
    }
    return df, summary


# ---------------------------------------------------------------------------
# 6) Dated run log
# ---------------------------------------------------------------------------

def write_run_info_xlsx(output_dir: str, wide_path: str, summary: dict) -> str:
    """Append one dated row per run to run_info.xlsx and refresh the latest snapshot."""
    from openpyxl import Workbook, load_workbook

    xlsx_path = os.path.join(output_dir, "run_info.xlsx")
    os.makedirs(output_dir, exist_ok=True)

    header = [
        "Timestamp", "Output Dir", "Wide CSV", "Seed",
        "tmax_hours", "step_hours", "noise_level",
        "n_curves", "n_valid", "n_invalid",
    ]

    wb = load_workbook(xlsx_path) if os.path.exists(xlsx_path) else Workbook()

    if "RUNS" in wb.sheetnames:
        ws_runs = wb["RUNS"]
        if ws_runs.max_row == 1 and ws_runs.max_column == 1 and ws_runs["A1"].value is None:
            ws_runs.append(header)
    else:
        ws_runs = wb.create_sheet("RUNS")
        ws_runs.append(header)

    ts = datetime.now().isoformat(timespec="seconds")
    ws_runs.append([
        ts, os.path.abspath(output_dir), os.path.abspath(wide_path), summary["seed"],
        summary["tmax_hours"], summary["step_hours"], summary["noise_level"],
        summary["n_curves"], summary["n_valid"], summary["n_invalid"],
    ])

    if "INFO" in wb.sheetnames:
        del wb["INFO"]
    ws_info = wb.create_sheet("INFO")
    ws_info["A1"] = (
        f"Output: {os.path.abspath(output_dir)} | File: {os.path.basename(wide_path)} | "
        f"Seed: {summary['seed']} | Timestamp: {ts}"
    )
    ws_info["A3"] = "n_curves"; ws_info["B3"] = summary["n_curves"]
    ws_info["A4"] = "n_valid"; ws_info["B4"] = summary["n_valid"]
    ws_info["A5"] = "n_invalid"; ws_info["B5"] = summary["n_invalid"]

    if "PARAMS" in wb.sheetnames:
        del wb["PARAMS"]
    ws_params = wb.create_sheet("PARAMS")
    ws_params.append(["subtype", "count"])
    for k, v in {**summary["valid_counts"], **summary["invalid_counts"]}.items():
        ws_params.append([k, v])

    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"]["A1"].value is None:
        del wb["Sheet"]
    wb.save(xlsx_path)
    return xlsx_path


# ---------------------------------------------------------------------------
# 7) CLI
# ---------------------------------------------------------------------------

def main(argv=None) -> int:
    p = argparse.ArgumentParser(description="GrowthQA synthetic growth-curve generator (wide CSV).")
    p.add_argument("--seed", type=int, default=123, help="Random seed.")
    p.add_argument("--max-time", type=float, default=16.0, help="Maximum time in hours.")
    p.add_argument("--time-step", type=float, default=0.5, help="Sampling interval in hours.")
    p.add_argument("--noise-level", type=float, default=0.05, help="Gaussian measurement-noise stdev.")
    p.add_argument("--output-dir", type=str, default="./gen_data", help="Output directory.")
    p.add_argument("--file-stem", type=str, default="syn", help="File stem for the output CSV.")
    args = p.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")

    df, summary = generate_synthetic_wide_df(
        tmax_hours=args.max_time,
        step_hours=args.time_step,
        seed=args.seed,
        noise_level=args.noise_level,
        file_stem=args.file_stem,
    )

    os.makedirs(args.output_dir, exist_ok=True)
    wide_path = os.path.join(args.output_dir, f"timeseries_wide_{args.file_stem}.csv")
    df.to_csv(wide_path, index=False)
    write_run_info_xlsx(args.output_dir, wide_path, summary)

    logging.info("Wrote %d curves (%d valid, %d invalid) to %s",
                 summary["n_curves"], summary["n_valid"], summary["n_invalid"], wide_path)
    logging.info("Valid subtype counts: %s", summary["valid_counts"])
    logging.info("Invalid subtype counts: %s", summary["invalid_counts"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())